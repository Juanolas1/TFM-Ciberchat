import json
import csv
import os
import PyPDF2
import docx
import openpyxl

from concurrent.futures import ThreadPoolExecutor
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.conf import settings
from .models import Chat, Message, Attachment


from .rag_langchain import index_attachment_chunks, retrieve_chunks
from langchain_google_genai import ChatGoogleGenerativeAI

# Tengo este modelo aquí a modo de mockup para que el chat esté funcional ya que el acceso a los sservers del irlab era temporal
llm = ChatGoogleGenerativeAI(
    model="gemini-2.0-flash-lite",
    google_api_key=settings.GOOGLE_API_KEY,
)


def history_to_prompt(history, max_turns=20) -> str:
    """Convierte los últimos turnos en texto plano para el prompt."""
    lines = []

    for m in history[-max_turns:]:
        role = "Usuario" if m.sender == "user" else "Asistente"
        lines.append(f"{role}: {m.content}")
    return "\n".join(lines)

def generate_chat_title(content: str, llm_instance) -> str:

    prompt = (
        "Basándote en el siguiente mensaje del usuario, genera un título descriptivo "
        "de MÁXIMO 5 palabras que resuma el tema de la conversación. "
        "Solo responde con el título, sin comillas ni explicaciones adicionales y en el mismo idioma que te pregunte el usuario puede ser español o inglés y está a continuación.\n\n"
        f"Mensaje: {content[:200]}"
    )
    try:
        title = llm_instance.invoke(prompt).content.strip()
        title = title.strip('"\'').strip()
        words = title.split()
        if len(words) > 5:
            title = " ".join(words[:5])
        return title if title else "Nuevo chat"
    except Exception as e:
        print(f"Error generando título: {e}")
        return "Nuevo chat"

def extract_text_from_file(file):
    """Devuelve texto plano del adjunto para indexarlo."""
    file_type = file.content_type
    file_name = file.name
    text = ""
    try:
        if "pdf" in file_type or file_name.lower().endswith(".pdf"):
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        elif "word" in file_type or file_name.lower().endswith((".docx", ".doc")):
            doc = docx.Document(file)
            for para in doc.paragraphs:
                text += para.text + "\n"
        elif "excel" in file_type or file_name.lower().endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(file)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"Hoja: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += ", ".join([str(c) if c is not None else "" for c in row]) + "\n"
        elif "csv" in file_type or file_name.lower().endswith(".csv"):
            decoded = file.read().decode("utf-8").splitlines()
            reader = csv.reader(decoded)
            for row in reader:
                text += ", ".join(row) + "\n"
        elif "text/" in file_type or file_name.lower().endswith(
            (".txt", ".md", ".json", ".xml", ".html", ".css", ".js")
        ):
            text = file.read().decode("utf-8")
        else:
            try:
                text = file.read().decode("utf-8")
            except UnicodeDecodeError:
                text = f"[No se pudo decodificar {file_name} como texto]"
        file.seek(0)  # reset puntero
        return text
    except Exception as e:
        print(f"Error extrayendo texto de {file_name}: {e}")
        file.seek(0)
        return f"[Error al procesar {file_name}: {e}]"

def call_llm_stream(prompt: str):

    try:
        stream = llm.stream(prompt)
        for chunk in stream:
            if chunk.content:
                yield chunk.content
    except Exception as e:
        print(f"Error en streaming: {e}")
        yield f"Error: {e}"

@require_http_methods(["GET", "POST"])
def chat_list_create(request):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No estás autenticado"}, status=401)
    
    if request.method == "GET":
        search = request.GET.get("search", "")
        if search:
            chats = Chat.objects.filter(
                Q(user=request.user)
                & (Q(title__icontains=search) | Q(messages__content__icontains=search))
            ).distinct()
        else:
            chats = Chat.objects.filter(user=request.user)
        data = [
            {
                "id": c.id,
                "title": c.title,
                "created_at": c.created_at.isoformat(),
                "updated_at": c.updated_at.isoformat(),
                "last_message": c.last_message,
                "unread_count": c.unread_count,
            }
            for c in chats
        ]
        return JsonResponse(data, safe=False)
    
    payload = json.loads(request.body or "{}")
    title = payload.get("title", "Nuevo chat")
    chat = Chat.objects.create(user=request.user, title=title)
    return JsonResponse(
        {
            "id": chat.id,
            "title": chat.title,
            "created_at": chat.created_at.isoformat(),
            "updated_at": chat.updated_at.isoformat(),
            "last_message": "",
            "unread_count": 0,
        },
        status=201,
    )

@require_http_methods(["GET", "PUT", "DELETE"])
def chat_detail(request, chat_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No estás autenticado"}, status=401)
    
    chat = get_object_or_404(Chat, id=chat_id, user=request.user)
    
    if request.method == "GET":
        return JsonResponse(
            {
                "id": chat.id,
                "title": chat.title,
                "created_at": chat.created_at.isoformat(),
                "updated_at": chat.updated_at.isoformat(),
                "last_message": chat.last_message,
                "unread_count": chat.unread_count,
            }
        )
    
    if request.method == "PUT":
        data = json.loads(request.body or "{}")
        if "title" in data:
            chat.title = data["title"]
            chat.save()
        return JsonResponse(
            {
                "id": chat.id,
                "title": chat.title,
                "created_at": chat.created_at.isoformat(),
                "updated_at": chat.updated_at.isoformat(),
                "last_message": chat.last_message,
                "unread_count": chat.unread_count,
            }
        )
    
    chat.delete()
    return JsonResponse({"detail": "Chat eliminado"})

@require_http_methods(["GET", "POST"])
def message_list_create(request, chat_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No estás autenticado"}, status=401)
    
    chat = get_object_or_404(Chat, id=chat_id, user=request.user)
    
    if request.method == "GET":
        chat.messages.filter(read=False, sender="assistant").update(read=True)
        term = request.GET.get("search", "")
        msgs = (
            chat.messages.filter(content__icontains=term)
            if term else chat.messages.all()
        )
        data = [
            {
                "id": m.id,
                "content": m.content,
                "sender": m.sender,
                "timestamp": m.timestamp.isoformat(),
                "read": m.read,
                "attachments": [
                    {
                        "id": a.id,
                        "filename": a.filename,
                        "file_type": a.file_type,
                        "file_size": a.file_size,
                        "url": request.build_absolute_uri(a.file.url),
                    }
                    for a in m.attachments.all()
                ],
            }
            for m in msgs
        ]
        return JsonResponse(data, safe=False)
    
    content = request.POST.get("content", "")
    user_msg = Message.objects.create(
        chat=chat, content=content, sender="user", read=True
    )
    
    attachments_json, file_texts = [], []
    for f in request.FILES.getlist("files"):
        txt = extract_text_from_file(f)
        attach = Attachment.objects.create(
            message=user_msg,
            file=f,
            filename=f.name,
            file_type=f.content_type,
            file_size=f.size,
            indexed=False,
        )
        index_attachment_chunks(request.user.id, chat.id, user_msg.id, attach.id, txt)
        attach.indexed = True
        attach.save()
        file_texts.append(f"Archivo: {f.name}\n\n{txt}\n")
        attachments_json.append(
            {
                "id": attach.id,
                "filename": attach.filename,
                "file_type": attach.file_type,
                "file_size": attach.file_size,
                "url": request.build_absolute_uri(attach.file.url),
            }
        )
    
    chat_history = list(chat.messages.order_by("timestamp"))
    hist_block = history_to_prompt(chat_history, max_turns=len(chat_history))
    
    ctx = retrieve_chunks(request.user.id, chat.id, content, k=20, include_user=False)
    ctx_block = "\n---\n".join(ctx)[:25_000]
    
    def generate_streaming_response():
        yield f"data: {json.dumps({'type': 'user_message', 'message': {'id': user_msg.id, 'content': user_msg.content, 'sender': 'user', 'timestamp': user_msg.timestamp.isoformat(), 'attachments': attachments_json, 'read': True}})}\n\n"
        
        assistant_msg = Message.objects.create(
            chat=chat, content="", sender="assistant", read=False
        )
        
        yield f"data: {json.dumps({'type': 'assistant_start', 'message_id': assistant_msg.id, 'timestamp': assistant_msg.timestamp.isoformat()})}\n\n"
        
        if ctx:
            prompt_rag = (
                "Eres un asistente de ciberseguridad. El usuario AUTORIZA que uses "
                "AUTORIZA que extraigas y muestres la información contenida en ellos. "
                "Esto NO viola su privacidad porque los datos fueron aportados "
                "voluntariamente.\n\n"
                "la información de los documentos adjuntos.\n\n"
                "Instrucciones:\n"
                "1. Usa la información entre <ctx></ctx> SOLO si es relevante.\n"
                "2. Responde de forma concisa, sin incluir <ctx>.\n"
                "3. Responde en el mismo idioma que la pregunta el usuario del final, puede ser español o inglés auqnue estas intrucciones están en español. \n"
                "4. Si la respuesta NO está en <ctx>, responde EXACTAMENTE: NEED_LLM.\n\n"
                f"Historial reciente:\n{hist_block}\n\n"
                f"<ctx>\n{ctx_block}\n</ctx>\n\n"
                f"Pregunta: {content}"
            )
            
            rag_answer = ""
            for chunk in call_llm_stream(prompt_rag):
                rag_answer += chunk
            
            if "NEED_LLM" in rag_answer.upper() or rag_answer.strip().upper() == "NEED_LLM":
                answer = ""
                for chunk in call_llm_stream(f"{hist_block}\n\n{content}"):
                    answer += chunk
                    yield f"data: {json.dumps({'type': 'chunk', 'content': chunk})}\n\n"
            else:
                answer = rag_answer
                chunk_size = 10
                for i in range(0, len(answer), chunk_size):
                    chunk = answer[i:i+chunk_size]
                    yield f"data: {json.dumps({'type': 'chunk', 'content': chunk})}\n\n"
        else:
            answer = ""
            for chunk in call_llm_stream(f"{hist_block}\n\n{content}"):
                answer += chunk
                yield f"data: {json.dumps({'type': 'chunk', 'content': chunk})}\n\n"
        
        assistant_msg.content = answer
        assistant_msg.save()
        
        user_messages_count = chat.messages.filter(sender="user").count()
        new_title = None
        if user_messages_count == 1:
            new_title = generate_chat_title(content, llm)
            chat.title = new_title
        
        chat.save()
        
        yield f"data: {json.dumps({'type': 'complete', 'chat_title_updated': new_title})}\n\n"
        yield "data: [DONE]\n\n"
    
    response = StreamingHttpResponse(
        generate_streaming_response(),
        content_type='text/event-stream'
    )
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response

@require_http_methods(["GET"])
def download_attachment(request, attachment_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No estás autenticado"}, status=401)
    
    attach = get_object_or_404(
        Attachment, id=attachment_id, message__chat__user=request.user
    )
    response = HttpResponse(attach.file, content_type=attach.file_type)
    response["Content-Disposition"] = f"attachment; filename=\"{attach.filename}\""
    return response

@require_http_methods(["GET"])
def search_messages(request):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No estás autenticado"}, status=401)
    
    q = request.GET.get("q", "")
    if not q:
        return JsonResponse({"detail": "Se requiere término de búsqueda"}, status=400)
    
    msgs = (
        Message.objects.filter(chat__user=request.user, content__icontains=q)
        .select_related("chat")
        .order_by("-timestamp")
    )
    grouped = {}
    for m in msgs:
        cid = m.chat.id
        grouped.setdefault(
            cid,
            {
                "chat_id": cid,
                "chat_title": m.chat.title,
                "messages": [],
            },
        )
        grouped[cid]["messages"].append(
            {
                "id": m.id,
                "content": m.content,
                "sender": m.sender,
                "timestamp": m.timestamp.isoformat(),
                "attachments": [
                    {
                        "id": a.id,
                        "filename": a.filename,
                        "file_type": a.file_type,
                        "file_size": a.file_size,
                        "url": request.build_absolute_uri(a.file.url),
                    }
                    for a in m.attachments.all()
                ],
            }
        )
    return JsonResponse(list(grouped.values()), safe=False)